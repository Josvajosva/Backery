from odoo import models, fields, _
from odoo.exceptions import UserError
import calendar
import io
import tempfile
import base64
import logging
import pytz
from datetime import datetime



_logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    raise UserError("openpyxl is required to import Excel files")

try:
    import csv
except ImportError:
    raise UserError("csv module not available")


class ImportAttendanceData(models.TransientModel):
    _name = "import.attendance.data"
    _description = "Attendance Card Swipe Import"

    file_select = fields.Binary(string="Select File", required=True)
    import_option = fields.Selection([
        ('xls', 'Excel File (XLS/XLSX)'),
        ('csv', 'CSV File')
    ], default='xls', required=True)

    def import_file(self):
        if not self.file_select:
            raise UserError(_("Please select a file"))

        if self.import_option == 'csv':
            return self._import_csv()
        return self._import_excel()

    def _import_csv(self):
        try:
            csv_data = base64.b64decode(self.file_select)
            data_file = io.StringIO(csv_data.decode("utf-8"))
            reader = list(csv.reader(data_file))
        except Exception as e:
            raise UserError(_("Invalid CSV file: %s") % e)

        if len(reader) < 2:
            raise UserError(_("CSV file has no data"))

        headers = [h.strip().lower().replace(' ', '_') for h in reader[0] if h]

        self._validate_headers(headers)

        success = error = 0

        for row_no, row in enumerate(reader[1:], 2):
            values = {
                headers[i]: str(row[i]).strip() if i < len(row) and row[i] else ''
                for i in range(len(headers))
            }

            try:
                self._create_attendance(values)
                success += 1
            except Exception as e:
                error += 1
                _logger.error("Row %s failed: %s", row_no, e)

        return self._result(success, error)



    def _import_excel(self):
        try:
            file_data = base64.b64decode(self.file_select)
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as fp:
                fp.write(file_data)
                file_path = fp.name

            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb.active
        except Exception as e:
            raise UserError(_("Invalid Excel file: %s") % e)

        headers = []
        header_row = 0


        for row_no, row in enumerate(sheet.iter_rows(values_only=True), 1):
            normalized = [
                str(c).strip().lower().replace(' ', '_') if c else ''
                for c in row
            ]
            if 'id' in normalized and 'date' in normalized:
                headers = normalized
                header_row = row_no
                break

        if not headers:
            raise UserError(_("Could not detect header row. Expected column 'ID'."))

        self._validate_headers(headers)

        success = error = 0

        for row_no, row in enumerate(
            sheet.iter_rows(min_row=header_row + 1, values_only=True),
            header_row + 1
        ):
            values = {}
            for i, key in enumerate(headers):
                if key:
                    values[key] = str(row[i]).strip() if i < len(row) and row[i] else ''

            try:
                self._create_attendance(values)
                success += 1
            except Exception as e:
                error += 1
                _logger.error("Row %s failed: %s", row_no, e)

        return self._result(success, error)

    def _parse_date(self, date_str):
        for fmt in ("%Y-%m-%d", "%d-%m-%Y","%Y-%m-%d %H:%M:%S", "%d/%m/%Y"):
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue
        raise UserError(
            _("Invalid date '%s'. Use YYYY-MM-DD or DD-MM-YYYY") % date_str
        )



    def _create_attendance(self, values):
        def pick(*keys):
            for k in keys:
                k = k.lower().replace(' ', '_')
                if values.get(k):
                    return values[k]
            return None

        emp_code = pick('id')
        date_str = pick('date')
        swipe = pick('record')

        _logger.info("ROW DATA => %s", values)

        if not emp_code:
            raise UserError(_("Employee ID is mandatory"))

        if not date_str:
            raise UserError(_("Date is mandatory"))

        if not swipe:
            raise UserError(_("Record column is mandatory and must contain Check-in;Check-out time"))

        if ';' not in swipe:
            raise UserError(
                _("Invalid Record value '%s'. Expected format HH:MM;HH:MM") % swipe
            )

        employee = self.env['hr.employee'].search([
            '|',
            ('barcode', '=', emp_code),
            ('identification_id', '=', emp_code)
        ], limit=1)

        if not employee:
            raise UserError(_("Employee not found: %s") % emp_code)

        try:
            att_date = self._parse_date(date_str)
        except Exception:
            raise UserError(_("Invalid date '%s'. Use YYYY-MM-DD") % date_str)

        times = swipe.split(';')

        user_tz = pytz.timezone(self.env.user.tz or 'UTC')

        # Check-in
        local_in = user_tz.localize(
            datetime.strptime(f"{att_date} {times[0]}", "%Y-%m-%d %H:%M")
        )
        check_in = local_in.astimezone(pytz.UTC).replace(tzinfo=None)

        # Check-out
        local_out = user_tz.localize(
            datetime.strptime(f"{att_date} {times[1]}", "%Y-%m-%d %H:%M")
        )
        check_out = local_out.astimezone(pytz.UTC).replace(tzinfo=None)

        location_name = pick('location')
        location_id = False

        if location_name:
            location = self.env["hr.work.location"].search([
                ("name", "=ilike", location_name.strip())
            ], limit=1)

            if location:
                location_id = location.id
            else:
                raise UserError(_("Location '%s' not found") % location_name)

        existing = self.env['hr.attendance'].search([
            ('employee_id', '=', employee.id),
            ('check_in', '=', check_in),
            ('check_out', '=', check_out),
            ('location_id', '=', location_id),
        ], limit=1)

        category_raw = pick('category')
        category_map = {
            'factory': 'factory',
            'general': 'general',
            'store': 'store',
            'factory shift': 'factory',
        }

        category_val = category_map.get(
            (category_raw or '').strip().lower()
        )

        vals = {
            'employee_id': employee.id,
            'check_in': check_in,
            'check_out': check_out,
            'category': category_val,
            'location_id': location_id,
        }

        if existing:
            existing.write(vals)
        else:
            self.env['hr.attendance'].create(vals)


    def _validate_headers(self, headers):
        required = {'id', 'date'}
        missing = required - set(headers)
        if missing:
            raise UserError(_("Missing required columns: %s") % ', '.join(missing))

    def _result(self, success, error):
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Import Result'),
                'message': _("Successfully imported: %s\nErrors: %s") % (success, error),
                'sticky': True,
                'type': 'success' if success else 'warning'
            }
        }

